#!/usr/bin/env python3

"""
Copyright 2024 Mainspring Research

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute,
sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS
IN THE SOFTWARE.
"""

import math
import sys

import numpy
import pandas as pd
from matplotlib import pyplot as plt
from dateutil.parser import parse
from openpyxl import Workbook
from openpyxl.styles import Font


def percOfTotal(value, total):
    return value / total


def currencyToFloat(row, currency, default=None):
    if default is not None and type(currency) is str and len(currency) == 0:
        if type(default) is str:
            return float(default.replace('$', '').replace(',', ''))
        else:
            return default
    elif currency is numpy.nan:
        return default
    else:
        if '--' in currency:
            print(f'{row["Symbol"]} - This equity has not yet been settled and the \n'
                  '"Total Gain/Loss Dollar" and/or "Cost Basis Total" is not yet available.\n'
                  'When this information becomes available the result will be more accurate.\n')
            return default
        return float(currency.replace('$', '').replace(',', ''))


def checkForDate(desc):
    dList = desc.split(' ')
    hasDate = False
    for d in dList:
        try:
            parse(d)
            hasDate = True
            break
        except ValueError:
            pass
    return hasDate


# Main
def main(fName, oDir):
    if fName[-4:] != '.csv':
        print('Input file must have a csv extension')
        exit(-1)
    fBaseName = fName[:len(fName)-4]
    i = fBaseName.rfind('/')
    if i != -1:
        fBaseName = fBaseName[i + 1:]

    oName = oDir
    if oName[len(oName) - 1] != '/':
        oName += '/'
    oName += fBaseName

    df = pd.read_csv(fName, index_col=False)

    # Sum pending activity
    pDF = df.loc[df['Symbol'] == 'Pending activity']
    pending = 0.0
    pDF.set_index('Symbol', inplace=True)

    for iX, iRow in pDF.iterrows():
        i = iRow['Last Price Change']
        amt = 0.0
        if type(i) is str:
            amt = currencyToFloat(iRow, i)
        elif type(i) is float:
            amt = i
        if math.isnan(amt):
            amt = 0.0
        amt = float(amt)
        if amt == 0.0:
            if type(iRow['Current Value']) is str:
                amt = currencyToFloat(iRow, iRow['Current Value'])
            elif type(iRow['Current Value']) is float:
                amt = iRow['Current Value']
        pending += float(amt)

    df = df[df['Description'] != 'BROKERAGELINK']
    df = df.filter(['Symbol', 'Description', 'Quantity', 'Current Value', 'Cost Basis Total'])
    df = df[df['Symbol'] != 'Pending activity']
    df = df[df['Current Value'].notnull()]
    df['Symbol'] = df['Symbol'].fillna('*CASH**')
    df['Quantity'] = df['Quantity'].fillna(1.0)
    df['Cost Basis Total'] = df['Cost Basis Total'].fillna(df['Current Value'])

    df['Current Value'] = df.apply(lambda rw: currencyToFloat(rw, rw['Current Value']), axis=1)
    df['Cost Basis Total'] = df.apply(lambda rw: currencyToFloat(rw, rw['Cost Basis Total'], rw['Current Value']), axis=1)

    df = df.reset_index(drop=True)
    df.loc[len(df.index)] = ['Pending**', 'Pending cash', 1.0, pending, pending]

    for index in df.index:
        if '**' in df.loc[index]['Symbol']:
            df.at[index, 'Symbol'] = '*CASH*'
            df.at[index, 'Description'] = 'Fixed Income'

    for index in df.index:
        if '%' in df.loc[index]['Description'] and checkForDate(df.loc[index]['Description']):
            if ' CD ' in df.loc[index]['Description']:
                df.at[index, 'Symbol'] = '*CDs'
            else:
                df.at[index, 'Symbol'] = '*Bonds'
            df.at[index, 'Description'] = 'Fixed Income'

    df = df.groupby(['Symbol', 'Description']).sum().reset_index()

    total = float(df.sum().loc['Current Value'])

    df.loc[df['Description'] == 'Fixed Income', 'Quantity'] = 1.0

    # Compute values in Python for pie chart, JSON, and sorting
    df.insert(3, 'Last Price', df['Current Value'] / df['Quantity'])
    df.insert(5, 'Average Cost Basis', df['Cost Basis Total'] / df['Quantity'])
    df['Gain-Loss'] = df['Current Value'] - df['Cost Basis Total']
    df['Gain-Loss %'] = df['Gain-Loss'] / df['Cost Basis Total']
    df['Perc of total'] = df.apply(lambda rw: percOfTotal(rw['Current Value'], total), axis=1)

    df = df.round(4)
    df = df.sort_values(by='Perc of total', ascending=False)
    dfFixed = df.query('Symbol.str.contains("*", regex=False)')
    dfNotFixed = df.query('not Symbol.str.contains("*", regex=False)')
    df = pd.concat([dfFixed, dfNotFixed], ignore_index=True)

    # --- Pie chart (unchanged) ---
    pieDict = df.to_dict(orient='index')
    labels = []
    perc = []
    for i in pieDict:
        labels.append(pieDict[i]['Symbol'])
        perc.append(pieDict[i]['Perc of total'])

    plt.figure(figsize=(15, 15))
    plt.rcParams.update({'font.size': 18})
    plt.pie(perc, labels=labels, autopct='{:.2f}%'.format)
    plt.savefig('{}.{}'.format(oName, 'png'), dpi='figure')

    # --- JSON (unchanged) ---
    jsonBuff = df.to_json(orient='records', indent=4)
    f = open('{}.json'.format(oName), 'w')
    f.write(jsonBuff)
    f.close()

    # --- Excel output with formulas ---
    # Column layout (1-indexed):
    #   A=1: Symbol
    #   B=2: Description
    #   C=3: Quantity
    #   D=4: Last Price           = E / C
    #   E=5: Current Value
    #   F=6: Average Cost Basis   = G / C
    #   G=7: Cost Basis Total
    #   H=8: Gain-Loss            = E - G
    #   I=9: Gain-Loss %          = H / G
    #   J=10: Perc of total       = E / total_cell

    headers = ['Symbol', 'Description', 'Quantity', 'Last Price', 'Current Value',
               'Average Cost Basis', 'Cost Basis Total', 'Gain-Loss', 'Gain-Loss %', 'Perc of total']

    num_data_rows = len(df)
    # Summary rows start after data
    total_row = num_data_rows + 2       # Excel row (1-indexed, +1 for header)
    stock_row = total_row + 1
    fixed_row = stock_row + 1

    # Determine which data rows are stocks vs fixed income for summary formulas
    stock_rows = []
    fixed_rows = []
    for idx, df_row in df.iterrows():
        excel_row = idx + 2  # +1 for header, +1 for 1-indexing
        if df_row['Description'] == 'Fixed Income':
            fixed_rows.append(excel_row)
        else:
            stock_rows.append(excel_row)

    def sum_formula(col_ref, target_rows):
        """Build a SUM of specific rows like =SUM(E2,E5,E8)"""
        if not target_rows:
            return 0
        return '=' + '+'.join(f'{col_ref}{rw}' for rw in target_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Investment Summary"

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data rows with formulas for calculated columns
    for df_idx, df_row in df.iterrows():
        xr = df_idx + 2  # Excel row

        ws.cell(row=xr, column=1, value=df_row['Symbol'])
        ws.cell(row=xr, column=2, value=df_row['Description'])
        ws.cell(row=xr, column=3, value=df_row['Quantity'])
        # D: Last Price = Current Value / Quantity
        ws.cell(row=xr, column=4).value = f'=IF(C{xr}=0,"",E{xr}/C{xr})'
        ws.cell(row=xr, column=5, value=df_row['Current Value'])
        # F: Average Cost Basis = Cost Basis Total / Quantity
        ws.cell(row=xr, column=6).value = f'=IF(C{xr}=0,"",G{xr}/C{xr})'
        ws.cell(row=xr, column=7, value=df_row['Cost Basis Total'])
        # H: Gain-Loss = Current Value - Cost Basis Total
        ws.cell(row=xr, column=8).value = f'=E{xr}-G{xr}'
        # I: Gain-Loss % = Gain-Loss / Cost Basis Total
        ws.cell(row=xr, column=9).value = f'=IF(G{xr}=0,"",H{xr}/G{xr})'
        # J: Perc of total = Current Value / grand total of Current Value
        # Reference the total row's Current Value cell
        ws.cell(row=xr, column=10).value = f'=IF(E{total_row}=0,"",E{xr}/E{total_row})'

    # --- Summary row: Total (not incl interest and dividends) ---
    last_data_row = num_data_rows + 1  # last Excel data row
    ws.cell(row=total_row, column=1, value='')
    ws.cell(row=total_row, column=2, value='Total (not incl interest and dividends)')
    ws.cell(row=total_row, column=3, value='')
    ws.cell(row=total_row, column=4, value='')
    ws.cell(row=total_row, column=5).value = f'=SUM(E2:E{last_data_row})'
    ws.cell(row=total_row, column=6, value='')
    ws.cell(row=total_row, column=7).value = f'=SUM(G2:G{last_data_row})'
    ws.cell(row=total_row, column=8).value = f'=E{total_row}-G{total_row}'
    ws.cell(row=total_row, column=9).value = f'=IF(G{total_row}=0,"",H{total_row}/G{total_row})'
    ws.cell(row=total_row, column=10, value='')

    # --- Summary row: Total stocks ---
    ws.cell(row=stock_row, column=1, value='')
    ws.cell(row=stock_row, column=2, value='Total stocks')
    ws.cell(row=stock_row, column=3, value='')
    ws.cell(row=stock_row, column=4, value='')
    ws.cell(row=stock_row, column=5).value = sum_formula('E', stock_rows)
    ws.cell(row=stock_row, column=6, value='')
    ws.cell(row=stock_row, column=7).value = sum_formula('G', stock_rows)
    ws.cell(row=stock_row, column=8).value = f'=E{stock_row}-G{stock_row}'
    ws.cell(row=stock_row, column=9).value = f'=IF(G{stock_row}=0,"",H{stock_row}/G{stock_row})'
    ws.cell(row=stock_row, column=10).value = f'=IF(E{total_row}=0,"",E{stock_row}/E{total_row})'

    # --- Summary row: Total fixed income ---
    ws.cell(row=fixed_row, column=1, value='')
    ws.cell(row=fixed_row, column=2, value='Total fixed income (not incl int. and div.)')
    ws.cell(row=fixed_row, column=3, value='')
    ws.cell(row=fixed_row, column=4, value='')
    ws.cell(row=fixed_row, column=5).value = f'=E{total_row}-E{stock_row}'
    ws.cell(row=fixed_row, column=6, value='')
    ws.cell(row=fixed_row, column=7).value = sum_formula('G', fixed_rows)
    ws.cell(row=fixed_row, column=8).value = f'=E{fixed_row}-G{fixed_row}'
    ws.cell(row=fixed_row, column=9).value = f'=IF(G{fixed_row}=0,"",H{fixed_row}/G{fixed_row})'
    ws.cell(row=fixed_row, column=10).value = f'=IF(E{total_row}=0,"",1-J{stock_row})'

    # --- Formatting ---
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.00%'

    for fmt_row in range(2, fixed_row + 1):
        for col in [4, 5, 6, 7, 8]:  # D, E, F, G, H
            ws.cell(row=fmt_row, column=col).number_format = currency_fmt
        for col in [9, 10]:  # I, J
            ws.cell(row=fmt_row, column=col).number_format = pct_fmt

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 12
    for cl in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[cl].width = 18
    for cl in ['I', 'J']:
        ws.column_dimensions[cl].width = 14

    # Bold summary rows
    for sr in [total_row, stock_row, fixed_row]:
        for c in range(1, 11):
            cell = ws.cell(row=sr, column=c)
            cell.font = Font(bold=True)

    wb.save('{}.xlsx'.format(oName))


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("USAGE: portfolioSummary.py input_file output_directory")
    main(sys.argv[1], sys.argv[2])
